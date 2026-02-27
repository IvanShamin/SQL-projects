#!/usr/bin/env python3
"""
unpivot.py — Flatten a 2-D Excel range of SQL INSERT fragments into a single column.

Usage:
    python unpivot.py input.xlsx --sheet inserts_composition --out flat_inserts.txt

Requirements:
    pip install openpyxl
"""

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: run  pip install openpyxl")

EMPTY_PATTERN = " ,)"  # rows containing this are empty compositions


def unpivot(wb_path: Path, sheet_name: str, out_path: Path, skip_empty: bool = True) -> None:
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        sys.exit(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows_written = 0
    rows_skipped = 0

    with out_path.open("w", encoding="utf-8") as fh:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                value = str(cell).strip()
                if not value:
                    continue
                if skip_empty and EMPTY_PATTERN in value:
                    rows_skipped += 1
                    continue
                fh.write(value + "\n")
                rows_written += 1

    print(f"✅  Done — {rows_written} fragments written, {rows_skipped} empty rows skipped.")
    print(f"    Output: {out_path.resolve()}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Unpivot Excel INSERT fragments to a flat file.")
    parser.add_argument("input",  type=Path, help="Path to the .xlsx workbook")
    parser.add_argument("--sheet", default="inserts_composition", help="Sheet name to read from")
    parser.add_argument("--out",   type=Path, default=Path("flat_inserts.txt"), help="Output file path")
    parser.add_argument("--keep-empty", action="store_true", help="Keep empty composition rows")
    args = parser.parse_args()

    if not args.input.exists():
        sys.exit(f"File not found: {args.input}")

    unpivot(args.input, args.sheet, args.out, skip_empty=not args.keep_empty)


if __name__ == "__main__":
    main()
